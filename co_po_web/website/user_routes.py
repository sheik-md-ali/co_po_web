# user_routes.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import sys
sys.path.append('D:/co_po_web')
from website.models import Subject, Section, InternalAssessment, CoAttainment, AssessmentInstance, College, SubjectList
from io import BytesIO
from website import db
from openpyxl import Workbook
import io
from .calulations import *



user_routes = Blueprint('user_routes', __name__)

@user_routes.route('/user/dashboard')
@login_required
def user_dashboard():
    return render_template('user/user_dashboard.html')


@user_routes.route('/user_subjects')
@login_required
def user_subjects():
    # Retrieve subjects for the current user by querying the database
    user_subjects = Subject.query.filter_by(user_id=current_user.id).all()

    return render_template('user/user_subjects.html', user_subjects=user_subjects)


@user_routes.route('/user/mapping_for_assessment', methods=['GET', 'POST'])
@login_required
def mapping_for_assessment():
    # Query distinct subject codes associated with the current user
    subject_codes = db.session.query(Subject.subject_code).filter_by(user_id=current_user.id).distinct().all()

    # Convert list of tuples to list of subject codes
    subject_codes = [code[0] for code in subject_codes]

    selected_subject_code = request.form.get('subject_code')
    selected_section_name = request.form.get('section')
    selected_assessment_instance_id = request.form.get('assessment_instance')

    sections = []
    assessment_instances = []
    assessment_details = None

    if selected_subject_code:
        # Query the sections based on the selected subject code and current user
        sections = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
            Subject.user_id == current_user.id,
            Subject.subject_code == selected_subject_code,
            Section.name == Subject.section
        ).all()


        if selected_section_name:
            # Query assessment instances based on the selected subject code and section name
            section = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
                Subject.user_id == current_user.id,
                Section.subject_code == selected_subject_code,
                Section.name == selected_section_name
            ).first()
            if section:
                assessment_instances = AssessmentInstance.query.filter_by(section_id=section.id).all()

                if selected_assessment_instance_id:
                    # Query assessment details if a specific instance is selected
                    assessment_instance = AssessmentInstance.query.get(selected_assessment_instance_id)
                    if assessment_instance:
                        assessment_details = {
                            'name': assessment_instance.name,
                            'max_marks': assessment_instance.assessment.max_marks
                        }

    return render_template('user/mapping_for_assessment.html', subjects=subject_codes, sections=sections,
                           assessment_instances=assessment_instances, assessment_details=assessment_details,
                           selected_subject_code=selected_subject_code, selected_section_name=selected_section_name,
                           selected_assessment_instance_id=selected_assessment_instance_id)


@user_routes.route('/user/save_or_submit_assessment', methods=['POST'])
@login_required
def save_or_submit_assessment():
    selected_subject_code = request.form.get('selected_subject_code')
    selected_section_name = request.form.get('selected_section_name')
    selected_assessment_instance_id = request.form.get('selected_assessment_instance_id')
    action = request.form.get('action')

    assessment_instance = AssessmentInstance.query.get(selected_assessment_instance_id)
    if not assessment_instance:
        flash('Assessment instance not found.', 'error')
        return redirect(url_for('user_routes.mapping_for_assessment'))

    if assessment_instance.status_for_mapping == 'submitted':
        flash('Assessment already submitted. Cannot modify further.', 'error')
        return redirect(url_for('user_routes.mapping_for_assessment'))

    mapping_dict = {}
    question_count = int(request.form.get('question_count', 0))


    for i in range(1, question_count + 1):
        max_marks = request.form.get(f'question{i}_maxMarks')
        co_selected = request.form.get(f'question{i}_co')

        if max_marks and co_selected:
            mapping_dict[f'Q{i}'] = {
                'maxMarks': max_marks,
                'co': co_selected
            }


    assessment_instance.mapping_dictionary = mapping_dict

    if action == 'submit':
        assessment_instance.status_for_mapping = 'submitted'
        flash('Assessment submitted successfully.', 'success')
    else:
        assessment_instance.status_for_mapping = 'saved'
        flash('Mapping saved successfully.', 'success')

    db.session.commit()

    return redirect(url_for('user_routes.mapping_for_assessment'))

        
@user_routes.route('/download/template.xlsx', methods=['GET'])
@login_required
def download_template():
    selected_assessment_instance_id = request.args.get('assessment_instance_id')
    
    if not selected_assessment_instance_id:
        flash('Please select an assessment instance.')
        return redirect(url_for('user_routes.upload_excel'))

    # Fetch assessment instance and associated mapping JSON
    assessment_instance = AssessmentInstance.query.get(selected_assessment_instance_id)
    if not assessment_instance:
        flash('Invalid assessment instance ID.')
        return redirect(url_for('user_routes.upload_excel'))
    
    mapping = assessment_instance.mapping_dictionary

    # Fetch necessary details for the template
    college = College.query.get(assessment_instance.college_id)
    subject = SubjectList.query.get(assessment_instance.subject_id)
    #user = User.query.get(subject.user_id)

    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Assessment Template'

    # Write header information
    sheet['A1'] = college.name
    sheet.merge_cells('A1:F1')

    # Add Year, Branch, and Semester in the second row
    sheet['A2'] = f"Year: {subject.year}  Branch: {subject.branch}  Semester: {subject.semester}"
    sheet.merge_cells('A2:F2')

    # Write subject code and subject name in the third row
    sheet['A3'] = f"SUBJECT CODE : {subject.subject_code}                     | SUBJECT NAME: {subject.subject_name}"
    sheet.merge_cells('A3:F3')

    # Write COs and Max Marks headers in the fourth and fifth rows
    sheet['C5'] = "CO's"
    sheet['C6'] = 'MAX MARKS'

    question_columns = list(mapping.keys())
    
    # Populate COs and Max Marks based on mapping
    for index, question in enumerate(question_columns, start=2):
        sheet.cell(row=5, column=index + 2).value = mapping[question].get('co', '')  # COs
        sheet.cell(row=6, column=index + 2).value = mapping[question].get('maxMarks', '')  # Max Marks

    # Write SNO, Student Name, REG NO, and Q1, Q2, ... headers in the seventh row
    sheet['A7'] = 'SNO'
    sheet['B7'] = 'STUDENT NAME'  # Add the Student Name column
    sheet['C7'] = 'REG NO'

    for index, _ in enumerate(question_columns, start=1):
        sheet.cell(row=7, column=index + 3).value = f"Q{index}"

    # Save workbook to a BytesIO object to send it as a downloadable file
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='assessment_template.xlsx'
    )


@user_routes.route('/user/upload_excel', methods=['GET', 'POST'])
@login_required
def upload_excel():
    subject_codes = db.session.query(Subject.subject_code).filter_by(user_id=current_user.id).distinct().all()
    subject_codes = [code[0] for code in subject_codes]

    selected_subject_code = request.form.get('subject_code')
    selected_section_name = request.form.get('section')
    selected_assessment_type = request.form.get('assessment_type')
    selected_assessment_instance_id = request.form.get('assessment_instance')

    sections = []
    assessment_instances = []
    assessment_details = None

    if selected_subject_code:
        sections = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
            Subject.user_id == current_user.id,
            Subject.subject_code == selected_subject_code
        ).all()

        if selected_section_name:
            section = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
                Subject.user_id == current_user.id,
                Section.subject_code == selected_subject_code,
                Section.name == selected_section_name
            ).first()
            if section and selected_assessment_type == 'assessments':
                assessment_instances = AssessmentInstance.query.filter_by(section_id=section.id).all()

                if selected_assessment_instance_id:
                    assessment_instance = AssessmentInstance.query.get(selected_assessment_instance_id)
                    if assessment_instance:
                        assessment_details = {
                            'name': assessment_instance.name,
                            'max_marks': assessment_instance.assessment.max_marks
                        }

    return render_template('user/upload_excel.html', subjects=subject_codes, sections=sections,
                           assessment_instances=assessment_instances, assessment_details=assessment_details,
                           selected_subject_code=selected_subject_code, selected_section_name=selected_section_name,
                           selected_assessment_instance_id=selected_assessment_instance_id,
                           selected_assessment_type=selected_assessment_type)

@user_routes.route('/user/upload_excel_file', methods=['POST'])
@login_required
def upload_excel_file():
    selected_assessment_instance_id = request.form.get('assessment_instance')
    selected_assessment_type = request.form.get('assessment_type')
    action = request.form.get('action')

    if selected_assessment_type not in ['assessments', 'semester_end_exam', 'course_exit_survey']:
        flash('Invalid assessment type.', 'danger')
        return redirect(url_for('user_routes.upload_excel'))

    if selected_assessment_type == 'assessments' and not selected_assessment_instance_id:
        flash('No assessment instance selected.', 'danger')
        return redirect(url_for('user_routes.upload_excel'))

    if 'assessmentfile' not in request.files:
        flash('No file part in the request.', 'danger')
        return redirect(url_for('user_routes.upload_excel'))

    file = request.files['assessmentfile']
    if file.filename == '':
        flash('No selected file.', 'danger')
        return redirect(url_for('user_routes.upload_excel'))

    if file:
        filename = secure_filename(file.filename)
        file_data = file.read()

        try:
            if selected_assessment_type == 'assessments':
                marks_data = read_excel_data(BytesIO(file_data))
                
                # Process the assessments data
                assessment_instance = AssessmentInstance.query.get(selected_assessment_instance_id)
                if not assessment_instance:
                    flash('Assessment instance not found.', 'danger')
                    return redirect(url_for('user_routes.upload_excel'))

                mapping_dict = assessment_instance.mapping_dictionary
                co_attainments = CoAttainment.query.filter_by(subject_code=assessment_instance.section.subject_code).all()
                
                if not co_attainments:
                    flash('No LOA data found for the selected subject code.', 'danger')
                    return redirect(url_for('user_routes.upload_excel'))

                loa_data = [co_attainment.loa_data for co_attainment in co_attainments]
                target_percs = co_attainments[0].target_perc  # Assuming target_perc is consistent across all COs for the subject

                individual_mapping, overall_mapping = create_mapping(mapping_dict, marks_data, loa_data, target_percs)

                individual_df = pd.DataFrame(marks_data)

                for co in mapping_dict.values():
                    co_key = co['co']
                    individual_df[f'{co_key}%'] = individual_df.apply(lambda row: individual_mapping.get(f"{row['REG NO']}_{co_key}", {}).get('percentage', ""), axis=1)
                    individual_df[f'{co_key}_LOA'] = individual_df.apply(lambda row: individual_mapping.get(f"{row['REG NO']}_{co_key}", {}).get('level_of_attainment', ""), axis=1)
                    individual_df[f'{co_key}_Target'] = individual_df.apply(lambda row: individual_mapping.get(f"{row['REG NO']}_{co_key}", {}).get('target_met', ""), axis=1)

                overall_row = [""] * len(individual_df.columns)
                overall_row[0] = "Overall"
                for co, data in overall_mapping.items():
                    co_index = individual_df.columns.get_loc(f'{co}%')
                    loa_index = individual_df.columns.get_loc(f'{co}_LOA')
                    target_index = individual_df.columns.get_loc(f'{co}_Target')
                    overall_row[co_index] = ""
                    overall_row[loa_index] = data['average_level_of_attainment']
                    overall_row[target_index] = data['target_met_count']['Y']

                overall_series = pd.Series(overall_row, index=individual_df.columns)
                individual_df = pd.concat([individual_df, pd.DataFrame([overall_series])], ignore_index=True)

                overall_co_df = pd.DataFrame(columns=['COs', 'Values', 'Target_count'])
                overall_co_df['COs'] = list(overall_mapping.keys())
                overall_co_df['Values'] = [data['average_level_of_attainment'] for data in overall_mapping.values()]
                overall_co_df['Target_count'] = [data['target_met_count']['Y'] for data in overall_mapping.values()]

                output = BytesIO()
                writer = pd.ExcelWriter(output, engine='xlsxwriter')
                workbook = writer.book
                worksheet = workbook.add_worksheet('Sheet1')

                college = College.query.get(assessment_instance.college_id)
                subject = SubjectList.query.get(assessment_instance.subject_id)
                worksheet.write('A1', college.name)
                worksheet.merge_range('A1:F1', college.name)
                worksheet.write('A2', f"Year: {subject.year}  Branch: {subject.branch}  Semester: {subject.semester}")
                worksheet.merge_range('A2:F2', f"Year: {subject.year}  Branch: {subject.branch}  Semester: {subject.semester}")
                worksheet.write('A3', f"SUBJECT CODE : {subject.subject_code}                     | SUBJECT NAME: {subject.subject_name}")
                worksheet.merge_range('A3:F3', f"SUBJECT CODE : {subject.subject_code}                     | SUBJECT NAME: {subject.subject_name}")

                worksheet.write('C5', "CO's")
                worksheet.write('C6', 'MAX MARKS')

                question_columns = list(mapping_dict.keys())

                for index, question in enumerate(question_columns, start=2):
                    worksheet.write(4, index + 1, mapping_dict[question].get('co', ''))  # COs
                    worksheet.write(5, index + 1, mapping_dict[question].get('maxMarks', ''))  # Max Marks

                individual_df.to_excel(writer, startrow=7, index=False, sheet_name='Sheet1')

                overall_co_df.to_excel(writer, startrow=len(individual_df) + 9, index=False, sheet_name='Sheet1')

                writer.close()
                updated_file_data = output.getvalue()

            elif selected_assessment_type == 'semester_end_exam':
                pass

            elif selected_assessment_type == 'course_exit_survey':
                pass

        except Exception as e:
            flash(f"Error processing Excel file: {e}", 'danger')
            return redirect(url_for('user_routes.upload_excel'))

        if selected_assessment_type == 'assessments':
            if action == 'submit' and assessment_instance.status_for_excel != 'submitted':
                assessment_instance.excel_file = updated_file_data
                assessment_instance.status_for_excel = 'submitted'
                db.session.commit()
                flash('Assessment file submitted successfully and cannot be modified further.', 'success')
            elif action == 'save' and assessment_instance.status_for_excel != 'submitted':
                assessment_instance.excel_file = updated_file_data
                assessment_instance.status_for_excel = 'saved'
                db.session.commit()
                flash('Assessment file saved successfully.', 'success')
            else:
                flash('Assessment file cannot be modified as it is already submitted.', 'danger')

        elif selected_assessment_type == 'semester_end_exam':
            # Handle save/submit logic if needed for semester end exam
            flash('Semester End Exam data uploaded successfully.', 'success')

        elif selected_assessment_type == 'course_exit_survey':
            # Handle save/submit logic if needed for course exit survey
            flash('Course Exit Survey data uploaded successfully.', 'success')

        return redirect(url_for('user_routes.upload_excel'))


@user_routes.route('/user/download_assessment', methods=['GET', 'POST'])
@login_required
def download_assessment():
    subject_codes = db.session.query(Subject.subject_code).filter_by(user_id=current_user.id).distinct().all()
    subject_codes = [code[0] for code in subject_codes]

    selected_subject_code = request.form.get('subject_code')
    selected_section_name = request.form.get('section')
    selected_assessment_instance_id = request.form.get('assessment_instance')

    sections = []
    assessment_instances = []
    assessment_details = None

    if selected_subject_code:
        sections = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
            Subject.user_id == current_user.id,
            Subject.subject_code == selected_subject_code
        ).all()

        if selected_section_name:
            section = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
                Subject.user_id == current_user.id,
                Section.subject_code == selected_subject_code,
                Section.name == selected_section_name
            ).first()
            if section:
                assessment_instances = AssessmentInstance.query.filter_by(section_id=section.id).all()

    if selected_subject_code and selected_section_name and selected_assessment_instance_id:
        assessment_instance = AssessmentInstance.query.get(selected_assessment_instance_id)
        if assessment_instance:
            assessment_details = {
                'name': assessment_instance.name,
                'max_marks': assessment_instance.assessment.max_marks
            }

    if request.method == 'POST' and 'download' in request.form:
        assessment_instance = AssessmentInstance.query.get(selected_assessment_instance_id)
        if not assessment_instance or not assessment_instance.excel_file:
            flash('Assessment instance not found or no Excel file available.', 'danger')
            return redirect(url_for('user_routes.download_assessment'))

        output = BytesIO(assessment_instance.excel_file)
        output.seek(0)
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'{assessment_instance.name}.xlsx')

    return render_template('user/download_assessment.html',
                           subjects=subject_codes,
                           sections=sections,
                           assessment_instances=assessment_instances,
                           selected_subject_code=selected_subject_code,
                           selected_section_name=selected_section_name,
                           selected_assessment_instance_id=selected_assessment_instance_id,
                           assessment_details=assessment_details)
from flask import flash

@user_routes.route('/user/internal_assessment_co', methods=['GET', 'POST'])
@login_required
def internal_assessment_co():
    subject_codes = db.session.query(Subject.subject_code).filter_by(user_id=current_user.id).distinct().all()
    subject_codes = [code[0] for code in subject_codes]

    selected_subject_code = request.form.get('subject_code')
    selected_section_name = request.form.get('section')
    calculation_method = request.form.get('calculation_method')
    
    co_values = {}
    target_counts = {}
    all_cos = set()
    assessment_names = {}  # Dictionary to hold assessment names

    sections = []

    if selected_subject_code:
        sections = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
            Subject.user_id == current_user.id,
            Subject.subject_code == selected_subject_code
        ).all()

        if selected_section_name:
            section = db.session.query(Section).join(Subject, Section.subject_code == Subject.subject_code).filter(
                Subject.user_id == current_user.id,
                Section.subject_code == selected_subject_code,
                Section.name == selected_section_name
            ).first()

            if section and calculation_method:
                subject = db.session.query(Subject).filter_by(subject_code=selected_subject_code).first()
                if subject:
                    internal_assessment = db.session.query(InternalAssessment).filter_by(
                        subject_id=subject.id,
                        section_id=section.id
                    ).first()

                    if internal_assessment:
                        if calculation_method == 'weightage':
                            co_values, target_counts, all_cos = calculate_co_weightage(internal_assessment.id)
                            internal_assessment.weightage_co_values = co_values  # Store weightage CO values
                        else:
                            co_values, target_counts, all_cos = calculate_co(internal_assessment.id)
                            internal_assessment.co_values = co_values  # Store normal CO values

                        # Fetch assessment names based on instance IDs in target_counts
                        instance_ids = list(target_counts.keys())
                        assessments = db.session.query(AssessmentInstance).filter(
                            AssessmentInstance.id.in_(instance_ids)
                        ).all()

                        for assessment in assessments:
                            assessment_names[assessment.id] = assessment.name

                        internal_assessment.target_counts = target_counts
                        db.session.commit()
                    else:
                        flash('No internal assessment is available for the selected section.', 'warning')
                else:
                    flash('Selected subject does not exist.', 'warning')
            else:
                if not section:
                    flash('Selected section does not exist.', 'warning')
                if not calculation_method:
                    flash('Please select a calculation method.', 'warning')
        else:
            flash('Please select a section.', 'warning')
    else:
        flash('Please select a subject.', 'warning')

    return render_template('user/internal_assessment_co.html',
                           subjects=subject_codes,
                           sections=sections,
                           selected_subject_code=selected_subject_code,
                           selected_section_name=selected_section_name,
                           co_values=co_values,
                           target_counts=target_counts,
                           all_cos=all_cos,
                           assessment_names=assessment_names,  # Pass the assessment names
                           calculation_method=calculation_method)